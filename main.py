import os
import time
import requests
import feedparser
from datetime import datetime
import traceback
import json
import atexit
from ntscraper import Nitter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from deep_translator import GoogleTranslator

# --- Kimlik bilgilerini .env dosyasından yükle ---
def _load_env():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ.setdefault(k.strip(), v.strip())
_load_env()

TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID   = os.environ.get("TELEGRAM_CHAT_ID", "")
GOOGLE_SHEET_URL   = os.environ.get("GOOGLE_SHEET_URL", "")

translation_cache = {}

def translate_text(text):
    if text in translation_cache:
        return translation_cache[text]

    for attempt in range(3):
        try:
            translated = GoogleTranslator(source='auto', target='tr').translate(text)
            translation_cache[text] = translated
            time.sleep(0.5)
            return translated
        except Exception:
            time.sleep(1)

    return text

def send_telegram_message(message, log_errors=True):
    """Telegram mesajı gönderir. Başarı/hata durumunu loglar."""
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("[UYARI] Telegram token veya chat_id tanımlı değil!")
        return False
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        data = {"chat_id": TELEGRAM_CHAT_ID, "text": message, "parse_mode": "HTML"}
        response = requests.post(url, data=data, timeout=10)
        if response.status_code == 200:
            print(f"[TELEGRAM ✓] Mesaj gönderildi.")
            return True
        else:
            err_msg = f"Telegram HTTP Hatası: {response.status_code} | {response.text[:200]}"
            print(f"[TELEGRAM ✗] {err_msg}")
            if log_errors:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                log_to_excel(
                    [timestamp, "-", "-", "-", "-", "-", "-", err_msg, "HATA", "-"],
                    status="Hata", report_type="Turkiye_Gundem"
                )
            return False
    except Exception as e:
        err_msg = f"Telegram Bağlantı Hatası: {e}"
        print(f"[TELEGRAM ✗] {err_msg}")
        if log_errors:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_to_excel(
                [timestamp, "-", "-", "-", "-", "-", "-", err_msg, "HATA", "-"],
                status="Hata", report_type="Turkiye_Gundem"
            )
        return False

# --- Ayarlar ve Konfigürasyon ---
RSS_URLS = [
    "https://www.haberturk.com/rss/ekonomi.xml",
    "https://www.ntv.com.tr/ekonomi.rss",
    "https://www.trthaber.com/ekonomi_articles.rss",
    "https://www.cnnturk.com/feed/rss/ekonomi/news"
]

KEYWORD_MAPPING = {
    "faiz": ["interest rate", "cbrt", "rate"],
    "tcmb": ["interest rate", "cbrt", "central bank"],
    "merkez bankası": ["interest rate", "cbrt", "central bank"],
    "enflasyon": ["inflation", "cpi"],
    "tüik": ["inflation", "cpi"],
    "dolar": ["usd", "try", "lira", "usd/try"],
    "kur": ["usd", "try", "lira", "exchange rate"],
    "seçim": ["election", "elections", "erdogan", "imamoglu", "yavas"],
    "türkiye": ["turkey", "turkish"],
    "türk": ["turkey", "turkish"]
}

NEWS_KEYWORDS = list(KEYWORD_MAPPING.keys())
TURKEY_POOL_KEYWORDS = ["turkey", "turkish", "erdogan", "lira", "cbrt", "istanbul", "ankara", "try"]
POLYMARKET_API_URL = "https://gamma-api.polymarket.com/markets"

MONTHS_TR = {
    1: "Ocak", 2: "Subat", 3: "Mart", 4: "Nisan",
    5: "Mayis", 6: "Haziran", 7: "Temmuz", 8: "Agustos",
    9: "Eylul", 10: "Ekim", 11: "Kasim", 12: "Aralik"
}

WEEKDAYS_TR = {
    0: "Pzt.", 1: "Sal.", 2: "Çarş.", 3: "Perş.",
    4: "Cum.", 5: "Cts.", 6: "Paz."
}

def get_current_excel_file_path():
    now = datetime.now()
    filename = f"Finansal_Radar_{now.year}.xlsx"
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

def get_current_sheet_name(report_type="Ekonomi"):
    # Report type'ları daha okunabilir yapalım
    mapping = {
        "Turkiye_Gundem": "Turkiye Gündemi",
        "Polymarket_Firsatlari": "Polymarket Fırsatları",
        "Ekonomi": "Ekonomi Takibi"
    }
    return mapping.get(report_type, report_type)

processed_news_links = set()
seen_analyses = set()

def analyze_news(title):
    title_lower = title.lower()
    if any(k in title_lower for k in ["altın", "xau", "ons", "gümüş", "emtia"]):
        return "ALTIN / EMTİA", "Örüntü: Altın ve emtia haberleri tarihsel olarak güvenli liman algısıyla talep yaratır. Aksiyon: Küresel belirsizlik durumunda XAU/USD long pozisyonları değerlendirilebilir."
    elif any(k in title_lower for k in ["dolar", "kur", "tl", "tcmb", "faiz", "enflasyon", "fed", "merkez bankası", "usd"]):
        return "DOLAR / TL DEĞER KAYBI", "Örüntü: Faiz ve enflasyon kararları yerel para birimi üzerinde doğrudan fiyatlanır. Aksiyon: Kararın beklenti dışı olmasına göre USD/TRY volatilitesi artabilir."
    elif any(k in title_lower for k in ["petrol", "brent", "enerji", "doğalgaz", "gaz", "opec"]):
        return "PETROL / ENERJİ", "Örüntü: Enerji arz/talep haberleri veya Orta Doğu gerilimleri Brent fiyatlarını hızlı fiyatlar. Aksiyon: Enerji şirketleri hisseleri ve petrol long izlenebilir."
    elif any(k in title_lower for k in ["kripto", "bitcoin", "btc", "ethereum", "web3", "sec", "coin"]):
        return "KRİPTO / WEB3", "Örüntü: Kripto regülasyon ve adaptasyon haberleri yüksek beta ile fiyatlanır. Aksiyon: İlgili coin/token özelinde hacim artışı takip edilmeli."
    elif any(k in title_lower for k in ["savaş", "kriz", "çatışma", "füze", "askeri", "abd", "rusya", "israil", "iran", "jeopolitik", "seçim", "siyaset"]):
        return "JEOPOLİTİK RİSK", "Örüntü: Jeopolitik risk artışı piyasalarda risk-off (riskten kaçış) yaratır. Aksiyon: Borsalarda satış, emtia ve tahvillerde alış baskısı."
    else:
        return "NÖTR", "Düşük Piyasa Etkisi"

def setup_sheet(ws):
    headers = [
        "Tarih / Saat", "Tetikleyici Kelime", "Haber Başlığı",
        "Haber Linki", "Polymarket Durumu", "Kontrat Adı",
        "Olasılık/Fiyat", "Sistem Raporu", "Fırsat Kategorisi", "Tarihsel Örüntü ve Etki Analizi"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    adjust_column_widths(ws)

def init_excel_for_today(report_type="Ekonomi"):
    file_path = get_current_excel_file_path()
    sheet_name = get_current_sheet_name(report_type)

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        setup_sheet(ws)
        wb.save(file_path)
    else:
        try:
            wb = load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
                setup_sheet(ws)
                wb.save(file_path)
        except Exception as e:
            print(f"Excel dosyası başlatılırken hata ({report_type}): {e}")

def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        if adjusted_width > 60:
            adjusted_width = 60
        ws.column_dimensions[column].width = adjusted_width

def log_to_google_sheet(row_data, report_type):
    """Veriyi anlık olarak Google Sheets'e gönderir."""
    if not GOOGLE_SHEET_URL:
        return
    
    try:
        sheet_name = get_current_sheet_name(report_type)
        payload = {
            "sheetName": sheet_name,
            "rowData": row_data
        }
        requests.post(GOOGLE_SHEET_URL, json=payload, timeout=10)
    except Exception as e:
        print(f"Google Sheets Log Hatası: {e}")

def log_to_excel(row_data, status="Bulunamadı", report_type="Ekonomi"):
    """
    row_data: [Tarih, Kelime, Başlık, Link, Durum, Kontrat, Fiyat, Rapor, Kategori, Analiz]
    """
    try:
        init_excel_for_today(report_type)
        file_path = get_current_excel_file_path()
        sheet_name = get_current_sheet_name(report_type)

        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        category = row_data[8]
        analysis_text = row_data[9]

        is_duplicate = False
        if status not in ["Sistem", "Hata"] and analysis_text not in ["-", "Düşük Piyasa Etkisi"]:
            if analysis_text in seen_analyses:
                is_duplicate = True
                row_data[9] = f"[TEKRAR] {analysis_text}"
            else:
                seen_analyses.add(analysis_text)

        kontrat_adi = row_data[5]
        existing_row = None
        if kontrat_adi and kontrat_adi != "-" and status in ["Bulundu", "Bulunamadı"]:
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row=row_idx, column=6).value == kontrat_adi:
                    existing_row = row_idx
                    break

        if existing_row:
            ws.cell(row=existing_row, column=7).value = row_data[6] # SADECE Olasılık/Fiyat güncellenir
            adjust_column_widths(ws)
            wb.save(file_path)
            return

        # Anlık Bulut Senkronizasyonu
        log_to_google_sheet(row_data, report_type)

        ws.append(row_data)
        current_row = ws.max_row

        link_str = str(row_data[3]) if row_data[3] is not None else "-"
        if link_str.startswith("http"):
            ws.cell(row=current_row, column=4).hyperlink = link_str
            ws.cell(row=current_row, column=4).font = Font(color="0563C1", underline="single")

        if status in ["Sistem", "Hata"] or category == "HATA":
            fill_color = "FFFFFF" # BEYAZ
        elif is_duplicate:
            fill_color = "40E0D0" # TURKUAZ
        elif status == "Bulunamadı":
            fill_color = "FFC7CE" # AÇIK KIRMIZI
        elif status == "Bulundu":
            if category in ["DOLAR / TL DEĞER KAYBI", "ARBITRAJ"]:
                fill_color = "00FF00" # Canlı YEŞİL
                if not is_duplicate:
                    if report_type == "Polymarket_Firsatlari":
                        msg = (
                            f"🟢 <b>YENİ POLYMARKET FIRSATI</b>\n"
                            f"━━━━━━━━━━━━━━━━━━━━\n"
                            f"<b>Kontrat:</b> {row_data[5]}\n"
                            f"<b>Olasılık/Fiyat:</b> {row_data[6]}\n"
                            f"<b>Zaman:</b> {row_data[0]}"
                        )
                    else:
                        haber_linki = f"\n<a href='{row_data[3]}'>📰 Habere Git</a>" if str(row_data[3]).startswith("http") else ""
                        msg = (
                            f"🟢 <b>YÜKSEK FIRSAT YAKALANDI</b>\n"
                            f"━━━━━━━━━━━━━━━━━━━━\n"
                            f"<b>Haber:</b> {row_data[2]}\n"
                            f"<b>Kategori:</b> {category}\n"
                            f"<b>Polymarket Kontratı:</b> {row_data[5]}\n"
                            f"<b>Olasılık/Fiyat:</b> {row_data[6]}\n"
                            f"<b>Tetikleyici:</b> {row_data[1]}"
                            f"{haber_linki}"
                        )
                    send_telegram_message(msg)
            elif category in ["ALTIN / EMTİA", "PETROL / ENERJİ"]:
                fill_color = "FFFF99" # Açık SARI
                if not is_duplicate and status == "Bulundu" and str(row_data[5]) != "-":
                    msg = (
                        f"🟡 <b>ORTA ÖNCELİKLİ FIRSAT</b>\n"
                        f"━━━━━━━━━━━━━━━━━━━━\n"
                        f"<b>Kategori:</b> {category}\n"
                        f"<b>Haber:</b> {row_data[2]}\n"
                        f"<b>Kontrat:</b> {row_data[5]}\n"
                        f"<b>Fiyat:</b> {row_data[6]}"
                    )
                    send_telegram_message(msg)
            elif category in ["KRİPTO / WEB3", "JEOPOLİTİK RİSK"]:
                fill_color = "FFD700" # Altın SARI
                if not is_duplicate and status == "Bulundu" and str(row_data[5]) != "-":
                    msg = (
                        f"🟠 <b>JEOPOLİTİK / KRİPTO ALARM</b>\n"
                        f"━━━━━━━━━━━━━━━━━━━━\n"
                        f"<b>Kategori:</b> {category}\n"
                        f"<b>Haber:</b> {row_data[2]}\n"
                        f"<b>Kontrat:</b> {row_data[5]}\n"
                        f"<b>Fiyat:</b> {row_data[6]}"
                    )
                    send_telegram_message(msg)
            else:
                fill_color = "FFFFFF" # BEYAZ
        else:
            fill_color = "FFFFFF"

        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for col_num in range(1, 11):
            cell = ws.cell(row=current_row, column=col_num)
            if col_num != 4:
                cell.fill = row_fill
            else:
                cell.fill = row_fill

        adjust_column_widths(ws)
        wb.save(file_path)
    except Exception as e:
        print(f"Excel'e yazarken hata oluştu: {e}")

def check_news():
    matched_news = []
    for url in RSS_URLS:
        try:
            feed = feedparser.parse(url)
            for entry in feed.entries:
                title = entry.title
                link = entry.link

                if link in processed_news_links:
                    continue

                title_lower = title.lower()
                for tr_keyword in NEWS_KEYWORDS:
                    if tr_keyword in title_lower:
                        matched_news.append((title, link, tr_keyword))
                        processed_news_links.add(link)
                        break
        except Exception as e:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_to_excel([timestamp, "-", "-", "-", "-", "-", "-", f"RSS Bağlantı Hatası: {e}", "HATA", "-"], status="Hata")

    return matched_news

twitter_fail_count = 0

def check_twitter_news():
    global twitter_fail_count
    matched_tweets = []
    try:
        scraper = Nitter(log_level=1)
        accounts = ["tcbestepe", "Merkez_Bankasi", "HMBakanligi"]
        for account in accounts:
            try:
                results = scraper.get_tweets(account, mode='user', number=3)
                if results and 'tweets' in results:
                    for tweet in results['tweets']:
                        text = tweet.get('text', '')
                        link = tweet.get('link', '')
                        if not link or link in processed_news_links:
                            continue

                        matched_tweets.append((f"[X] {account}: {text[:80]}...", link, "resmi gazete"))
                        processed_news_links.add(link)
            except Exception as e:
                print(f"Twitter hesap hatası ({account}): {e}")

        twitter_fail_count = 0
    except Exception as e:
        twitter_fail_count += 1
        if twitter_fail_count == 1:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_to_excel([timestamp, "-", "-", "-", "-", "-", "-", f"Twitter Nitter Hatası: {e}", "HATA", "-"], status="Hata", report_type="Turkiye_Gundem")
        print(f"Twitter Modül Hatası: {e}")

    return matched_tweets

def get_turkey_markets_pool():
    turkey_pool = []
    try:
        offset = 0
        limit = 100
        while offset < 3000:
            params = {"active": "true", "closed": "false", "limit": limit, "offset": offset}
            response = requests.get(POLYMARKET_API_URL, params=params, timeout=15)

            if response.status_code == 200:
                markets = response.json()
                if not markets:
                    break

                for market in markets:
                    if market.get("closed") or not market.get("active"):
                        continue

                    question = market.get("question", "")
                    question_lower = question.lower()

                    if any(k in question_lower for k in TURKEY_POOL_KEYWORDS):
                        yes_price = None
                        no_price = None

                        outcomes = market.get("outcomes", [])
                        prices = market.get("outcomePrices", [])

                        if isinstance(outcomes, str):
                            try:
                                outcomes = json.loads(outcomes)
                            except:
                                outcomes = []

                        if isinstance(prices, str):
                            try:
                                prices = json.loads(prices)
                            except:
                                prices = []

                        if outcomes and prices and len(outcomes) == len(prices):
                            for i, outcome in enumerate(outcomes):
                                if outcome == "Yes":
                                    yes_price = prices[i]
                                elif outcome == "No":
                                    no_price = prices[i]
                        else:
                            tokens = market.get("tokens", [])
                            for token in tokens:
                                if token.get("outcome") == "Yes":
                                    yes_price = token.get("price")
                                elif token.get("outcome") == "No":
                                    no_price = token.get("price")

                        invalid_prices = (None, "N/A", "None", "Null", "0", "0.0")
                        if yes_price in invalid_prices or no_price in invalid_prices:
                            continue

                        # Format prices with $ sign
                        yes_price_str = f"${float(yes_price):.3f}" if str(yes_price).replace('.', '', 1).isdigit() else f"${yes_price}"
                        no_price_str = f"${float(no_price):.3f}" if str(no_price).replace('.', '', 1).isdigit() else f"${no_price}"

                        turkey_pool.append({
                            "question": question,
                            "question_lower": question_lower,
                            "yes_price": yes_price_str,
                            "no_price": no_price_str
                        })
                offset += limit
            else:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                log_to_excel([timestamp, "-", "-", "-", "-", "-", "-", f"Polymarket HTTP Hatası: {response.status_code}", "HATA", "-"], status="Hata")
                break

    except requests.exceptions.RequestException as e:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_to_excel([timestamp, "-", "-", "-", "-", "-", "-", f"Polymarket Bağlantı Hatası: {e}", "HATA", "-"], status="Hata")
    except Exception as e:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_to_excel([timestamp, "-", "-", "-", "-", "-", "-", f"Polymarket Beklenmeyen Hata: {e}", "HATA", "-"], status="Hata")

    return turkey_pool

def match_markets(turkey_pool, tr_keyword, title):
    matched_markets = []
    en_keywords = KEYWORD_MAPPING.get(tr_keyword, [])

    # Haberin İngilizce çevirisi (bağlam kontrolü için)
    translated_title = translate_text(title).lower()

    for market in turkey_pool:
        question_lower = market["question_lower"]

        # 1. Anahtar kelime kesişimi
        has_keyword = any(en_k in question_lower for en_k in en_keywords)

        if has_keyword:
            # 2. Katı Bağlam Filtresi (Strict Semantic Context)
            # Sorudaki anahtar kelimelerin (stop words hariç) haberin çevirisinde geçip geçmediği kontrol edilir.
            words = set(question_lower.replace("?", "").replace(",", "").replace(".", "").split())
            stop_words = {"will", "be", "by", "in", "to", "the", "a", "an", "is", "are", "of", "and", "or", "for", "on", "at", "it", "this", "that"}
            important_words = words - stop_words

            # Haberle soru arasındaki organik kesişim (4 harften uzun önemli kelimeler)
            overlap = [w for w in important_words if len(w) > 3 and w in translated_title]

            strong_context = False
            # Eğer keyword doğrudan hem soruda hem haberde geçiyorsa bu çok güçlü bir bağdır
            for en_k in en_keywords:
                if en_k in question_lower and en_k in translated_title:
                    strong_context = True
                    break

            # Eğer güçlü bir bağ varsa veya önemli bağlam kelimeleri örtüşüyorsa eşleşmeyi onayla
            if strong_context or len(overlap) >= 1:
                matched_markets.append(market)

    return matched_markets

def run_live_api_test():
    print("--- CANLI API TESTİ BAŞLATILIYOR ---")
    try:
        params = {"active": "true", "closed": "false", "title": "turkey", "limit": 10}
        response = requests.get(POLYMARKET_API_URL, params=params, timeout=15)
        if response.status_code == 200:
            markets = response.json()
            if markets:
                market = markets[0]
                question = market.get("question", "Bilinmeyen Kontrat")

                yes_price = None
                no_price = None

                outcomes = market.get("outcomes", [])
                prices = market.get("outcomePrices", [])

                if isinstance(outcomes, str):
                    try:
                        outcomes = json.loads(outcomes)
                    except:
                        outcomes = []
                if isinstance(prices, str):
                    try:
                        prices = json.loads(prices)
                    except:
                        prices = []

                if outcomes and prices and len(outcomes) == len(prices):
                    for i, outcome in enumerate(outcomes):
                        if outcome == "Yes":
                            yes_price = prices[i]
                        elif outcome == "No":
                            no_price = prices[i]
                else:
                    tokens = market.get("tokens", [])
                    for token in tokens:
                        if token.get("outcome") == "Yes":
                            yes_price = token.get("price")
                        elif token.get("outcome") == "No":
                            no_price = token.get("price")

                yes_price_str = f"${float(yes_price):.3f}" if yes_price is not None and str(yes_price).replace('.', '', 1).isdigit() else f"${yes_price}"
                no_price_str = f"${float(no_price):.3f}" if no_price is not None and str(no_price).replace('.', '', 1).isdigit() else f"${no_price}"
                price_str = f"Yes: {yes_price_str} | No: {no_price_str}"
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                translated_question = translate_text(question)
                if translated_question and translated_question != question:
                    display_question = f"{question} ({translated_question})"
                else:
                    display_question = question

                log_to_excel([
                    timestamp,
                    "CANLI-API-TEST",
                    "(Polymarket Gerçek Veri Bağlantı Testi)",
                    "-",
                    "Bulundu",
                    display_question,
                    price_str,
                    "Başarılı",
                    "TEST",
                    "API bağlantısı başarılı, gerçek veri çekildi."
                ], status="Bulundu", report_type="Polymarket_Firsatlari")
                print("Canlı API testi başarılı. Excel'e yazıldı.")
            else:
                print("Canlı API testinde 'Turkey' başlıklı aktif kontrat bulunamadı.")
        else:
            print(f"Canlı API testinde HTTP hatası: {response.status_code}")
    except Exception as e:
        print(f"Canlı API testinde hata oluştu: {e}")

def main():
    start_time = datetime.now()
    ok = send_telegram_message(
        "🚀 <b>BOT BAŞLATILDI</b>\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "✅ Telegram bildirim köprüsü aktif\n"
        "📡 RSS + Twitter + Polymarket radarı çalışıyor\n"
        f"🕐 Başlangıç: {start_time.strftime('%d.%m.%Y %H:%M:%S')}",
        log_errors=False
    )
    if not ok:
        print("[KRİTİK UYARI] Telegram bağlantısı başlangıçta BAŞARISIZ! Token/Chat ID kontrol et.")

    # Bot kapandığında Telegram'a bildirim gönder
    def on_exit():
        send_telegram_message(
            "⛔ <b>BOT DURDU</b>\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            f"🕐 Başlangıç: {start_time.strftime('%d.%m.%Y %H:%M:%S')}\n"
            f"🕑 Kapanış : {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n"
            "⚠️ Bot yeniden başlatılmalı!",
            log_errors=False
        )
    atexit.register(on_exit)

    print("Sistem mimarisi (2 Excel + Akıllı Bildirim) son haline getirildi, operasyon tamamen aktif.")
    init_excel_for_today("Turkiye_Gundem")
    init_excel_for_today("Polymarket_Firsatlari")

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_to_excel([timestamp, "SİSTEM BAŞLATILDI", "Bot V7 - Gece Yarısı Geçişi Düzeltmesi.", "-", "-", "-", "-", "Başarılı", "NÖTR", "-"], status="Sistem", report_type="Turkiye_Gundem")
    log_to_excel([timestamp, "SİSTEM BAŞLATILDI", "Bot V7 - Gece Yarısı Geçişi Düzeltmesi.", "-", "-", "-", "-", "Başarılı", "NÖTR", "-"], status="Sistem", report_type="Polymarket_Firsatlari")

    # Kontrat havuzu: soru → son fiyat (fiyat değişirse tekrar bildirim gider)
    seen_polymarket_contracts = {}  # question -> last_price_str
    last_heartbeat_hour = -1        # saatlik kalp atışı takibi
    last_sheet_name = get_current_sheet_name()  # gece yarısı geçişi takibi
    run_live_api_test()

    while True:
        try:
            # --- Gece yarısı sayfa geçişi kontrolü ---
            current_sheet_name = get_current_sheet_name()
            if current_sheet_name != last_sheet_name:
                print(f"[GECEYARıSı] Yeni gün: {current_sheet_name} sayfası oluşturuluyor...")
                init_excel_for_today("Turkiye_Gundem")
                init_excel_for_today("Polymarket_Firsatlari")
                seen_polymarket_contracts.clear()  # Yeni günde tüm kontratları yeniden logla
                last_sheet_name = current_sheet_name
                send_telegram_message(
                    f"🌅 <b>YENİ GÜN BAŞLADI</b>\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"📅 Tarih: {current_sheet_name}\n"
                    f"📊 Excel sayfası oluşturuldu, izleme devam ediyor.",
                    log_errors=False
                )

            print(f"[{datetime.now().strftime('%H:%M:%S')}] Yeni haberler, tweetler ve Polymarket kontrol ediliyor...")
            new_relevant_news = check_news()
            twitter_news = check_twitter_news()
            all_news = new_relevant_news + twitter_news

            turkey_pool = get_turkey_markets_pool()
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # DOSYA 1: Polymarket Fırsatları
            if turkey_pool:
                for market in turkey_pool:
                    question = market['question']
                    price_str = f"Yes: {market['yes_price']} | No: {market['no_price']}"

                    is_new_contract = question not in seen_polymarket_contracts
                    price_changed = not is_new_contract and seen_polymarket_contracts.get(question) != price_str

                    if not is_new_contract and not price_changed:
                        continue  # Ne yeni ne de fiyat değişmiş → atla

                    seen_polymarket_contracts[question] = price_str

                    translated_question = translate_text(question)
                    display_question = f"{question} ({translated_question})" if translated_question and translated_question != question else question

                    rapor_notu = "Yeni kontrat tespit edildi." if is_new_contract else f"Fiyat güncellendi: {price_str}"

                    log_to_excel([
                        timestamp, "-", "Polymarket Yeni Kontrat", "-",
                        "Bulundu", display_question, price_str, "Başarılı",
                        "ARBITRAJ", rapor_notu
                    ], status="Bulundu", report_type="Polymarket_Firsatlari")

            # DOSYA 2: Turkiye Gundem
            if all_news:
                print(f"{len(all_news)} adet ilgili haber/tweet bulundu. Excel'e yazılıyor.")
                for title, link, tr_keyword in all_news:
                    report_type = "Turkiye_Gundem"

                    matched_poly_markets = match_markets(turkey_pool, tr_keyword, title)

                    category, analysis = analyze_news(title)
                    if "resmi gazete" in title.lower() or "resmi gazete" in tr_keyword.lower():
                        category = "DOLAR / TL DEĞER KAYBI" # Yüksek fırsat (Yeşil) olması için
                        analysis = "Örüntü: Resmi Gazete kararları doğrudan piyasa etkisine sahiptir."

                    if matched_poly_markets:
                        for market in matched_poly_markets:
                            question = market['question']
                            translated_question = translate_text(question)
                            if translated_question and translated_question != question:
                                display_question = f"{question} ({translated_question})"
                            else:
                                display_question = question

                            price_str = f"Yes: {market['yes_price']} | No: {market['no_price']}"
                            log_to_excel([
                                timestamp, tr_keyword.upper(), title, link,
                                "Bulundu", display_question, price_str, "Başarılı",
                                category, analysis
                            ], status="Bulundu", report_type=report_type)
                    else:
                        log_to_excel([
                            timestamp, tr_keyword.upper(), title, link,
                            "Bulunamadı", "-", "-", "Başarılı",
                            category, analysis
                        ], status="Bulunamadı", report_type=report_type)
            else:
                log_to_excel([timestamp, "-", "Sistem tarandı, yeni haber bulunamadı.", "-", "-", "-", "-", "Çalışıyor", "NÖTR", "Düşük Piyasa Etkisi"], status="Sistem", report_type="Turkiye_Gundem")

        except Exception as e:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            tb = traceback.format_exc()
            log_to_excel([timestamp, "-", "-", "-", "-", "-", "-", f"Ana Döngü Hatası: {e}", "HATA", "-"], status="Hata", report_type="Turkiye_Gundem")
            log_to_excel([timestamp, "-", "-", "-", "-", "-", "-", f"Ana Döngü Hatası: {e}", "HATA", "-"], status="Hata", report_type="Polymarket_Firsatlari")
            print(f"Hata yakalandı ve loglandı: {e}\n{tb}")
            send_telegram_message(
                f"🔴 <b>BOT KRİTİK HATA</b>\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"<b>Hata:</b> {str(e)[:300]}\n"
                f"<b>Zaman:</b> {timestamp}"
            )

        # --- Saatlik Kalp Atışı Bildirimi ---
        current_hour = datetime.now().hour
        if current_hour != last_heartbeat_hour:
            last_heartbeat_hour = current_hour
            now_str = datetime.now().strftime('%d.%m.%Y %H:%M')
            send_telegram_message(
                f"💓 <b>SİSTEM AKTİF</b> — {now_str}\n"
                f"Toplam taranan kontrat: {len(seen_polymarket_contracts)}\n"
                f"Sonraki kontrol: 1 dakika sonra.",
                log_errors=False
            )

        time.sleep(60)

if __name__ == "__main__":
    main()